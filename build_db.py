import openpyxl, csv, os, json

BASE_DIR = r'd:\CoreThree\TRIO\questions assets'
OUTPUT_FILE = r'd:\CoreThree\TRIO\app\js\questions.json'

DB = {
    'round1': [],
    'round2': [],
    'round3': []
}

def clean(s):
    if s is None: return ''
    return str(s).strip()

def parse_round1(fpath, cat):
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    # Find header row
    start_r = 1
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and (row[0] == '#' or row[1] == 'السؤال' or row[0] == 'السؤال'):
            start_r = i + 1
            break
            
    for row in ws.iter_rows(min_row=start_r+1, values_only=True):
        if not row or not row[0] or not row[1]: continue
        question = clean(row[1])
        answer = clean(row[2]) if len(row) > 2 else ''
        notes = clean(row[3]) if len(row) > 3 else ''
        
        # skip structural rows
        if question == '' or question == 'السؤال' or question.startswith('المصدر'): continue
        
        ar_text = question
        if answer:
            ar_text += f'\n(الإجابة: {answer})'
            
        DB['round1'].append({
            'category': cat,
            'ar': ar_text,
            'en': ar_text # for english toggle, using arabic as default if not translated
        })

def parse_round2(fpath, cat):
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    start_r = 1
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and row[0] == '#':
            start_r = i + 1
            break
            
    for row in ws.iter_rows(min_row=start_r+1, values_only=True):
        if not row or not row[0] or not row[1]: continue
        q = clean(row[1])
        answers = []
        # Answers are in cols 2-8
        for j in range(2, min(9, len(row))):
            a = clean(row[j])
            if a:
                # Remove trophies / medals
                a = a.replace('🏆','').replace('🥈','').replace('🥉','').strip()
                answers.append(a)
                
        if len(answers) == 7:
            # Generate points: 40 down to 10
            points = [40, 35, 30, 25, 20, 15, 10]
            DB['round2'].append({
                'category': cat,
                'ar': q,
                'en': q,
                'answers_ar': answers,
                'answers_en': answers,
                'points': points
            })

def parse_round3(fpath, cat, is_csv=False):
    if is_csv:
        with open(fpath, encoding='utf-8', errors='ignore') as fh:
            reader = csv.reader(fh)
            for i, row in enumerate(reader):
                if i == 0 or not row or len(row) < 2: continue
                q = clean(row[0])
                a = clean(row[1])
                text = f'{q}\n(Ans: {a})'
                DB['round3'].append({
                    'category': cat,
                    'ar': text,
                    'en': text
                })
    else:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        start_r = 1
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if row and (row[0] == '#' or row[1] == 'السؤال' or row[0] == 'السؤال'):
                start_r = i + 1
                break
                
        for row in ws.iter_rows(min_row=start_r+1, values_only=True):
            if not row or not row[0]: continue
            
            # Check if column 0 holds the question or col 1
            if len(row) > 1 and clean(row[1]):
                q = clean(row[1])
                a = clean(row[2]) if len(row) > 2 else ''
            else:
                q = clean(row[0])
                a = clean(row[1]) if len(row) > 1 else ''
                
            if q == '' or q == 'السؤال': continue
            
            ar_text = q
            if a: ar_text += f'\n(الإجابة: {a})'
                
            DB['round3'].append({
                'category': cat,
                'ar': ar_text,
                'en': ar_text
            })

# -- Round 1 --
r1_base = os.path.join(BASE_DIR, 'Turn-Based questions')
parse_round1(os.path.join(r1_base, 'geo multi.xlsx'), 'geography')
parse_round1(os.path.join(r1_base, 'history multi.xlsx'), 'history')
parse_round1(os.path.join(r1_base, 'islamic multi v3.xlsx'), 'islamic')
parse_round1(os.path.join(r1_base, 'تحدي الدوره كره القدم.xlsx'), 'football')
parse_round1(os.path.join(r1_base, 'تحدي الدور معلومات عامة.xlsx'), 'general')
parse_round1(os.path.join(r1_base, 'أعمال فنانين تحدي الدور.xlsx'), 'art')

# -- Round 2 --
r2_base = os.path.join(BASE_DIR, 'family feud questions')
# the 7 means there's 7 answers, mapping it to 'general' since it's mixed
parse_round2(os.path.join(r2_base, 'family feud 7.xlsx'), 'general')

# -- Round 3 --
r3_base = os.path.join(BASE_DIR, 'fast based questions')
parse_round3(os.path.join(r3_base, 'art.xlsx'), 'art')
parse_round3(os.path.join(r3_base, 'football_quiz_questions.csv'), 'football', True)
parse_round3(os.path.join(r3_base, 'geography easy.xlsx'), 'geography')
parse_round3(os.path.join(r3_base, 'islamic quiz (1).xlsx'), 'islamic')
parse_round3(os.path.join(r3_base, 'world history easy.xlsx'), 'history')

# Print stats
print(f"R1 Questions: {len(DB['round1'])}")
print(f"R2 Feuds: {len(DB['round2'])}")
print(f"R3 Questions: {len(DB['round3'])}")

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    json.dump(DB, f, ensure_ascii=False, indent=2)

print(f"Successfully generated {OUTPUT_FILE}!")
